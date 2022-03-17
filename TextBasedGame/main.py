from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import re


workbook = load_workbook("Platsnamn textspel.xlsx")
sheet = workbook['Rum']
GameMap = workbook['Karta']
savefile = workbook['Sparande']



#sheet.append(["ID",'Room','Room Description', 'Paths'])



room_description = 1
directions = 2


def goThroughSheet(thesheet):
    # goes through a sheet and puts every value into a 2d matrix (I think i can use that word),grid also works
    main_map=[] # the matrix that is going to be returned
    for row in thesheet:
        column_rooms=[]
        for column in row:
            column_rooms.append(str(column.value))#adds a cell from sheet row to the lists row
        main_map.append(column_rooms)
    return main_map


def look(room_value):#takes a list of the roomtype
    #the description of the roo
    print(room_value[room_description])

def save(x,y):# här kommer spar funktionen
    savefile['A1']= x
    savefile['A2']= y
    workbook.save('Platsnamn textspel.xlsx')
    print('Your file has been saved')



def move(x, y, direction, valids):#här kommer rörelse funktionen
    direction = direction.capitalize()
    latitude = {'East': 1, 'West': -1}
    longitude = {'North': -1, 'South':1}

    if direction in valids and direction in latitude:
        x += latitude[direction]
    if direction in valids and direction in longitude:
        y += longitude[direction]
    else:
        print('enter valid direction')

    """ 
    if direction == 'North':
        y -= 1
    elif direction == 'South':
        y += 1
    elif direction == 'East':
        x += 1
    elif direction == 'West':
        x -= 1
    else:
        print('Enter a valid direction')
    """
    return x, y


#print(goThroughSheet(GameMap))

def main():
    x = 0
    y = 1


    x = int(savefile['A1'].value)
    y = int(savefile['A2'].value)

    running = True
    room_IDs = goThroughSheet(sheet)
    main_map = goThroughSheet(GameMap)
    player_position = main_map[y][x]
    current_room_type = room_IDs[int(float(player_position))]
    print("let's begin")
    while running:
        valid_directions = current_room_type[directions].split(" ")

        print('you can move', end="")
        for show_direction in valid_directions:
            print(", "+str(show_direction), end="")
        print('')
        a = input()
        if a == 'look':
            look(current_room_type)


        if a == 'save':
            save(x,y)

        if re.search('move ', a):
            text = a
            text = re.sub('move ', '', text)

            x, y = move(x, y, text, valid_directions)
            print(text)
        if a == 'quit':
           running= False


        player_position = main_map[y][x]
        current_room_type = room_IDs[int(float(player_position))]
        #print(player_position)

        #print(str(x) + ' ' + str(y))
        #print(str(current_room_type))

        #for element in main_map:
        #    print(element)

main()
#sheet.append(["1",'hallway','you see a brightly lit hallway infront of you', 'North,South'])

#workbook.save("hello_world.xlsx")

